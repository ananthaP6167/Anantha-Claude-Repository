"""
CUSIP Transformation Macro
--------------------------
Reads an Excel file, applies =LEFT(A,3) on the Cusip column,
writes the result to a new 'transformed cusip' column, and saves output.

Usage:  python3 cusip_transform_macro.py <input_file> <output_file>
"""

import sys
import openpyxl
from openpyxl.styles import Font


def apply_transform(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    last_row = ws.max_row

    # Write header in B1
    ws["B1"] = "transformed cusip"
    if ws["A1"].font:
        ws["B1"].font = Font(name="Arial", bold=True, size=ws["A1"].font.size or 11)

    # Apply =LEFT(A,3) formula for every data row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=2, value=f"=LEFT(A{r},3)")

    # Auto-width column B
    ws.column_dimensions["B"].width = max(ws.column_dimensions["A"].width or 14, 18)

    wb.save(output_path)
    return last_row - 1  # rows processed


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 cusip_transform_macro.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    n = apply_transform(sys.argv[1], sys.argv[2])
    print(f"Processed {n} rows -> {sys.argv[2]}")
